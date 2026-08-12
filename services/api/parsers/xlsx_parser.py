"""
XLSX 字幕表格解析器
直接复用 ScriptGrid 的 parsers/xlsx_parser.py 逻辑
"""
import logging

logger = logging.getLogger(__name__)

EXCEL_HEADERS = ["序号", "开始时间", "结束时间", "字幕内容"]


def parse_xlsx(file_path):
    """
    解析 .xlsx 字幕表格文件
    :return: [[序号, 开始时间, 结束时间, 字幕内容], ...]
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=file_path, read_only=True)
        ws = wb.active
        header_row = [cell.value for cell in next(ws.iter_rows())]
        expected_len = len(EXCEL_HEADERS)
        if len(header_row) < expected_len:
            raise ValueError(f"Excel 文件表头不正确。期望: {EXCEL_HEADERS}, 实际: {header_row}")
        if header_row[:expected_len] != EXCEL_HEADERS:
            raise ValueError(f"Excel 文件表头不正确。期望: {EXCEL_HEADERS}, 实际: {header_row[:expected_len]}")
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(cell is not None for cell in row):
                continue
            if len(row) < 4:
                continue
            index, start_time, end_time, text = row[0], row[1], row[2], row[3]
            data.append([str(index), str(start_time), str(end_time), str(text) if text else ""])
        return data
    except ValueError:
        raise
    except Exception as e:
        raise ValueError(f"解析 Excel 文件 '{file_path}' 时出错: {e}") from e
